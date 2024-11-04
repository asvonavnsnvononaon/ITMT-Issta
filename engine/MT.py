import sys
import os
from imagecorruptions import get_corruption_names
from imagecorruptions import corrupt
import cv2
from openpyxl import Workbook
import engine.Add_object as Add_object
#from engine.MT_resize import MT_diffusion
from diffusers import FluxInpaintPipeline
#from Diffusion_tools.diffusion.TCD.demo_inpainting import prompt
from tqdm import tqdm
import gc
import os
current_dir = os.getcwd()
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)
import numpy as np
from PIL import Image, ImageFilter
import json
import os
import pandas as pd
import engine.data_process as data_process
import random
import shutil
from engine.RIFE import interpolate
import torch
from diffusers import StableDiffusionInstructPix2PixPipeline, EulerAncestralDiscreteScheduler
import engine.test_ADS as test_ADS
import torchvision.transforms as transforms
from diffusers import StableDiffusionXLImg2ImgPipeline
from diffusers.utils import load_image
import math
from PIL import Image, ImageOps
def get_informations(args):
    if args.dataset == "udacity":
        datasets = ["HMB1", "HMB2", "HMB4", "HMB5", "HMB6"]
    else:
        datasets = ["camera_lidar-20180810150607", "camera_lidar-20190401121727", "camera_lidar-20190401145936"]
    dir_1 = os.path.join(args.data_file, "ADS_data", args.dataset)#改变大小的图片的位置
    dir_2 = os.path.join(args.data_file, "ADS_data", args.dataset, "OneFormer")#分割图像的位置，尺寸是对应原来的图像
    dir_3 = os.path.join(args.data_file, "ADS_data", args.dataset, "langsam","road")
    dir_4 = os.path.join(args.data_file, "ADS_data", args.dataset, "langsam", "roadside")
    return dir_1,dir_2,dir_3,dir_4,datasets

def get_informations_1(args):
    if args.dataset == "A2D2":
        dir_1 = os.path.join(args.data_file, "ADS_data",args.dataset)  # 改变大小的图片的位置
        dir_2 = os.path.join(args.data_file, "ADS_data",args.dataset, "OneFormer")
        random_dataset =  "camera_lidar-20180810150607"
    else:
        dir_1 = os.path.join(args.data_file, "ADS_data",args.dataset)  # 改变大小的图片的位置
        dir_2 = os.path.join(args.data_file, "ADS_data",args.dataset, "OneFormer")
        random_dataset = "HMB5"
    return dir_1,dir_2,random_dataset
def find_test_images(args):
    dir_1,dir_2,dir_3,dir_4, datasets = get_informations_1(args)
    random_dataset = random.choice(datasets)
    save_dir = os.path.join(args.data_file, "MT_results","original")
    save_dir_2 = os.path.join(args.data_file, "MT_results", "follow_up")
    save_dir_1 = os.path.join(args.data_file, "MT_results", "mask")
    save_dir_3 = os.path.join(args.data_file, "MT_results", "onroad_mask")
    save_dir_4 = os.path.join(args.data_file, "MT_results", "roadside_mask")
    for dir in [save_dir, save_dir_1, save_dir_2,save_dir_3,save_dir_4]:
        shutil.rmtree(dir, ignore_errors=True)
        data_process.Check_file(dir)
    ######################################################################################
    rs = pd.read_excel(os.path.join(dir_1, random_dataset, "matched_data.xlsx"), header=0)
    rs = rs.iloc[1:-1]
    max_start = len(rs) - args.MT_image_num
    if args.dataset == "udacity":
        start_index = random.randint(0, max_start)
    else:
        start_index = random.randint(2656, max_start)
    time_series = args.pre_series
    number_test_images = args.MT_image_num-args.MT_image_num%time_series

    segmentation_maps = []
    data_list = []
    selected = rs[start_index:start_index + number_test_images]
    selected = selected.reset_index(drop=True)
    for idx in range(args.MT_image_num):
        image = os.path.basename(selected['Image File'][idx])
        image_path = os.path.join(dir_1,random_dataset,"center",image)
        mask_path = os.path.join(dir_2,random_dataset,"center",image)
        onroad_mask_path = os.path.join(dir_3,random_dataset,"center",image)
        roadside_mask_path = os.path.join(dir_4, random_dataset, "center", image)
        img =Image.open(image_path)
        mask = Image.open(mask_path)
        onroad_mask = Image.open(onroad_mask_path)
        roadside_mask = Image.open(roadside_mask_path)
        save_path = os.path.join(save_dir, image)
        save_path_1= os.path.join(save_dir_1, image)
        save_path_2 = os.path.join(save_dir_2, image)
        save_path_3 = os.path.join(save_dir_3, image)
        save_path_4 = os.path.join(save_dir_4, image)
        img.save(save_path)
        mask.save(save_path_1)
        onroad_mask.save(save_path_3)
        roadside_mask.save(save_path_4)
        data = {
            'Image File': save_path,
            'Mask File': save_path_1,
            'result File': save_path_2,
            "onroad mask":save_path_3,
            "roadside mask":save_path_4,
            'Steering Angle': selected['Steering Angle'][idx],
            'Vehicle Speed': selected['Vehicle Speed'][idx],
        }
        data_list.append(data)
    ####################################################################################
    with open(os.path.join(args.data_file, "MT_results", "info.json"), 'w', encoding='utf-8') as f:
        json.dump(data_list, f, ensure_ascii=False, indent=4)
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Data"
    headers = ["Image File", "Mask File", 'result File',"onroad mask","roadside mask", "Steering Angle", "Vehicle Speed"]
    ws.append(headers)
    for data in data_list:
        ws.append([data['Image File'], data['Mask File'], data['result File'],data['onroad mask'],data['roadside mask'],
                   data['Steering Angle'],data['Vehicle Speed']])
    excel_filename = os.path.join(args.data_file, "MT_results", "matched_data.xlsx")
    wb.save(excel_filename)


def find_test_images_1(args):
    data_list = []
    save_dir = os.path.join(args.data_file, "results","original")
    save_dir_2 = os.path.join(args.data_file, "results", "follow_up")
    save_dir_1 = os.path.join(args.data_file, "results", "mask")
    for dir in [save_dir, save_dir_1, save_dir_2]:
        shutil.rmtree(dir, ignore_errors=True)
        data_process.Check_file(dir)
    ######################################################################################
    args.dataset = "A2D2"
    dir_1, dir_2, random_dataset = get_informations_1(args)
    rs = pd.read_excel(os.path.join(dir_1, random_dataset, "matched_data.xlsx"), header=0)
    rs = rs.iloc[1:-1]

    #A2D2
    start_index = 2656#5214 2656
    time_series = 25
    number_test_images = (len(rs) - start_index) - (len(rs) - start_index) % time_series
    selected = rs[start_index:start_index + number_test_images]
    selected = selected.reset_index(drop=True)

    for idx in range(number_test_images):
        image = os.path.basename(selected['Image File'][idx])
        image_path = os.path.join(dir_1,random_dataset,"center",image)
        mask_path = os.path.join(dir_2,random_dataset,"center",image)
        img =Image.open(image_path)
        mask = Image.open(mask_path)
        save_path = os.path.join(save_dir, image)
        save_path_1= os.path.join(save_dir_1, image)
        save_path_2 = os.path.join(save_dir_2, image)
        img.save(save_path)
        mask.save(save_path_1)
        data = {
            'Image File': save_path,
            'Mask File': save_path_1,
            'result File': save_path_2,
            'Steering Angle': selected['Steering Angle'][idx],
            'Vehicle Speed': selected['Vehicle Speed'][idx],
        }
        data_list.append(data)
    ######################################################################################
    args.dataset = "udacity"
    dir_1, dir_2, random_dataset = get_informations_1(args)
    rs = pd.read_excel(os.path.join(dir_1, random_dataset, "matched_data.xlsx"), header=0)
    rs = rs.iloc[1:-1]

    # A2D2
    start_index = 0  # 5214 2656
    time_series = 25
    number_test_images = (len(rs) - start_index) - (len(rs) - start_index) % time_series
    selected = rs[start_index:start_index + number_test_images]
    selected = selected.reset_index(drop=True)

    for idx in range(number_test_images):
        image = os.path.basename(selected['Image File'][idx])
        image_path = os.path.join(dir_1, random_dataset, "center", image)
        mask_path = os.path.join(dir_2, random_dataset, "center", image)
        img = Image.open(image_path)
        mask = Image.open(mask_path)
        save_path = os.path.join(save_dir, image)
        save_path_1 = os.path.join(save_dir_1, image)
        save_path_2 = os.path.join(save_dir_2, image)
        img.save(save_path)
        mask.save(save_path_1)
        data = {
            'Image File': save_path,
            'Mask File': save_path_1,
            'result File': save_path_2,
            'Steering Angle': selected['Steering Angle'][idx],
            'Vehicle Speed': selected['Vehicle Speed'][idx],
        }
        data_list.append(data)


    with open(os.path.join(args.data_file, f"results", "info.json"), 'w', encoding='utf-8') as f:
        json.dump(data_list, f, ensure_ascii=False, indent=4)
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Data"
    headers = ["Image File", "Mask File", 'result File', "Steering Angle", "Vehicle Speed"]
    ws.append(headers)
    for data in data_list:
        ws.append([data['Image File'], data['Mask File'], data['result File'],
                   data['Steering Angle'],data['Vehicle Speed']])
    excel_filename = os.path.join(args.data_file, f"results", "matched_data.xlsx")
    wb.save(excel_filename)


def find_test_images_2(args, save_path, rule_number):

    data_list = []

    save_dir = os.path.join(save_path, "original")
    save_dir_2 = os.path.join(save_path, "follow_up")
    save_dir_1 = os.path.join(save_path, "mask")
    diffusion_path = os.path.join(args.data_file, "results", "follow_up", "Exp_2", str(rule_number),"videos")
    save_diffusion = os.path.join(save_path, "follow_up","0")
    for dir in [save_dir, save_dir_1, save_dir_2,save_diffusion]:
        data_process.Check_file(dir)
    if rule_number < 8:

        args.dataset = "A2D2"
        dir_1, dir_2, random_dataset = get_informations_1(args)
        random_dataset = "camera_lidar-20180810150607"
        rs = pd.read_excel(os.path.join(dir_1, random_dataset, "matched_data.xlsx"), header=0)
        rs = rs.iloc[1:-1]
        # A2D2
        start_index = 2656  # 5214 2656
        time_series = 25
        image_num = 15*25
        max_start = len(rs) - image_num-5
        start = start_index+rule_number*image_num
        if start>max_start:
            start = max_start
    else:
        args.dataset = "udacity"
        dir_1, dir_2, random_dataset = get_informations_1(args)
        random_dataset = "HMB5"
        rs = pd.read_excel(os.path.join(dir_1, random_dataset, "matched_data.xlsx"), header=0)
        rs = rs.iloc[1:-1]
        start_index = 0  # 5214 2656
        time_series = 25
        image_num = 15 * 25
        max_start = len(rs) - args.MT_image_num-100
        start = start_index + (rule_number-8) * image_num
        if start > max_start:
            start = max_start
    selected = rs[start:start + image_num]
    selected = selected.reset_index(drop=True)
    for idx in range(image_num):
        image = os.path.basename(selected['Image File'][idx])
        image_path = os.path.join(dir_1, random_dataset, "center", image)
        mask_path = os.path.join(dir_2, random_dataset, "center", image)
        diffusion_path_new = os.path.join(diffusion_path,image)
        diffusion_image= Image.open(diffusion_path_new)
        diffusion_image = diffusion_image.resize((320, 160), Image.LANCZOS)

        diffusion_image.save(os.path.join(save_diffusion, image))
        img = Image.open(image_path)
        mask = Image.open(mask_path)
        save_path_ = os.path.join(save_dir, image)
        save_path_1 = os.path.join(save_dir_1, image)
        save_path_2 = os.path.join(save_dir_2, image)
        img = img.resize((320, 160), Image.LANCZOS)
        mask = mask.resize((320, 160), Image.NEAREST)
        img.save(save_path_)
        mask.save(save_path_1)
        data = {
            'Image File': save_path_,
            'Mask File': save_path_1,
            'result File': save_path_2,
            'Steering Angle': selected['Steering Angle'][idx],
            'Vehicle Speed': selected['Vehicle Speed'][idx],
        }
        data_list.append(data)
    with open(os.path.join(save_path, "info.json"), 'w', encoding='utf-8') as f:
        json.dump(data_list, f, ensure_ascii=False, indent=4)
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Data"
    headers = ["Image File", "Mask File", 'result File', "Steering Angle", "Vehicle Speed"]
    ws.append(headers)
    for data in data_list:
        ws.append([data['Image File'], data['Mask File'], data['result File'],
                   data['Steering Angle'],data['Vehicle Speed']])
    excel_filename = os.path.join(save_path, "matched_data.xlsx")
    wb.save(excel_filename)

def find_test_images_1000(args, save_path):

    data_list = []

    save_dir = os.path.join(save_path, "original")
    save_dir_2 = os.path.join(save_path, "follow_up")
    save_dir_1 = os.path.join(save_path, "mask")
    rule_number = 20
    for ii in range(rule_number):
        #diffusion_path = os.path.join(args.data_file, "results", "follow_up", "Exp_2", str(ii), "videos")
        save_diffusion = os.path.join(save_path, "MRs", str(ii))
        data_process.Check_file(save_diffusion)


    for dir in [save_dir, save_dir_1, save_dir_2]:
        data_process.Check_file(dir)

    args.dataset = "A2D2"
    dir_1, dir_2, random_dataset = get_informations_1(args)
    random_dataset = "camera_lidar-20180810150607"
    rs = pd.read_excel(os.path.join(dir_1, random_dataset, "matched_data.xlsx"), header=0)
    rs = rs.iloc[1:-1]
    start_index = 2656  # 5214 2656
    time_series = 25
    number_test_images = (len(rs) - start_index) - (len(rs) - start_index) % time_series
    rs = rs[start_index:start_index + number_test_images]
    args.dataset = "udacity"
    dir_1, dir_2, random_dataset = get_informations_1(args)
    random_dataset = "HMB5"
    rs_1 = pd.read_excel(os.path.join(dir_1, random_dataset, "matched_data.xlsx"), header=0)
    rs_1 = rs_1.iloc[1:-1]
    start_index = 0  # 5214 2656
    number_test_images = (len(rs_1) - start_index) - (len(rs_1) - start_index) % time_series
    rs_1 = rs_1[start_index:start_index + number_test_images]

    rs = pd.concat([rs, rs_1], ignore_index=True)
    segment_size=25
    total_segments = len(rs) // segment_size
    selected_indices = random.sample(range(total_segments),40)
    sampled_segments = []
    data_list=[]
    for idx in selected_indices:
        if idx < 102:
            args.dataset = "A2D2"
            random_dataset = "camera_lidar-20180810150607"
        else:
            args.dataset = "udacity"
            random_dataset = "HMB5"
        dir_1 = os.path.join(args.data_file, "ADS_data", args.dataset)
        dir_2 = os.path.join(args.data_file, "ADS_data", args.dataset, "OneFormer")

        start = idx * segment_size
        for i in tqdm(range(segment_size)):
            image = os.path.basename(rs['Image File'][start+i])
            image_path = os.path.join(dir_1, random_dataset, "center", image)
            mask_path = os.path.join(dir_2, random_dataset, "center", image)
            img = Image.open(image_path)
            mask = Image.open(mask_path)
            save_path_ = os.path.join(save_dir, image)
            save_path_1 = os.path.join(save_dir_1, image)
            save_path_2 = os.path.join(save_dir_2, image)
            for ii in range(rule_number):
                diffusion_path = os.path.join(args.data_file, "results", "follow_up", "Exp_2", str(ii), "videos")
                save_diffusion = os.path.join(save_path, "MRs", str(ii))
                diffusion_path_new = os.path.join(diffusion_path, image)
                diffusion_image = Image.open(diffusion_path_new)
                diffusion_image = diffusion_image.resize((320, 160), Image.LANCZOS)
                diffusion_image.save(os.path.join(save_diffusion, image))


            img = img.resize((320, 160), Image.LANCZOS)
            mask = mask.resize((320, 160), Image.NEAREST)
            img.save(save_path_)
            mask.save(save_path_1)
            data = {
                'Image File': save_path_,
                'Mask File': save_path_1,
                'result File': save_path_2,
                'Steering Angle': rs['Steering Angle'][idx+i],
                'Vehicle Speed': rs['Vehicle Speed'][idx+i],
            }
            data_list.append(data)
    with open(os.path.join(save_path, "info.json"), 'w', encoding='utf-8') as f:
        json.dump(data_list, f, ensure_ascii=False, indent=4)
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Data"
    headers = ["Image File", "Mask File", 'result File', "Steering Angle", "Vehicle Speed"]
    ws.append(headers)
    for data in data_list:
        ws.append([data['Image File'], data['Mask File'], data['result File'],
                   data['Steering Angle'], data['Vehicle Speed']])
    excel_filename = os.path.join(save_path, "matched_data.xlsx")
    wb.save(excel_filename)


#################################
#exp2#
import contextlib
import io
import os
from PIL import Image
import cv2
import functools
def suppress_output(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            return func(*args, **kwargs)
    return wrapper

@suppress_output
def EXP_Pix2Pix(pix2pix,data_lists,input_dir,mask_dir,output_dir,start_index,end_index,diffusion_prompt):
    type = 1
    if type==1:
        for idx in range(start_index, end_index):
            if idx % 2 == 0 or idx == end_index-1 or idx==start_index:
                image_path = os.path.join(input_dir, os.path.basename(data_lists[idx]['Image File']))
                image = Image.open(image_path)
                processed_image = pix2pix(diffusion_prompt, image)
                processed_image.save(os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File'])))
        for idx in range(start_index, end_index):
            if idx % 2 == 1 and idx!=end_index-1 and idx!=start_index:
                image_path = os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File']))
                image_path_0 = os.path.join(output_dir, os.path.basename(data_lists[idx - 1]['Image File']))
                image_path_1 = os.path.join(output_dir, os.path.basename(data_lists[idx + 1]['Image File']))
                fig_new = interpolate(image_path_0, image_path_1)
                cv2.imwrite(image_path, fig_new)
    else:
        for idx in range(start_index, end_index):
            image_path = os.path.join(input_dir, os.path.basename(data_lists[idx]['Image File']))
            image = Image.open(image_path)
            processed_image = pix2pix(diffusion_prompt, image)
            processed_image.save(os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File'])))

def EXP_img2img(pix2pix,data_lists,input_dir,mask_dir,output_dir,start_index,end_index,diffusion_prompt):

    for idx in range(start_index, end_index):
        image_path = os.path.join(input_dir, os.path.basename(data_lists[idx]['Image File']))

        init_image = load_image(image_path).convert("RGB")
        prompt ="photorealistic"+diffusion_prompt+ "detailed, 8k"
        image = pix2pix(prompt, image=init_image,use_karras_sigmas=True,).images
        image[0].save(os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File'])))



def EXP_diffusion(inpainter, data_lists, input_dir, mask_dir, output_dir, start_index, end_index, diffusion_prompt):
    image_path = os.path.join(input_dir, os.path.basename(data_lists[start_index]['Image File']))
    mask_path = os.path.join(mask_dir, os.path.basename(data_lists[start_index]['Image File']))
    image = inpainter(image_path, mask_path, prompt=diffusion_prompt)
    image.save(os.path.join(output_dir,"images", os.path.basename(data_lists[start_index]['Image File'])))


def EXP_vista(args,vista_gen, data_lists, input_dir, mask_dir, output_dir, start_index, end_index, diffusion_prompt):
    image_path = os.path.join(output_dir,"images", os.path.basename(data_lists[start_index]['Image File']))
    speed = data_lists[start_index]['Vehicle Speed']
    speed = speed * 5
    generated_images = vista_gen(image_path, speed)
    for cont in range(0, args.pre_series):
        image_path_ = os.path.join(output_dir,"videos", os.path.basename(data_lists[start_index + cont]['Image File']))
        to_pil = transforms.ToPILImage()
        img = to_pil(generated_images[cont])
        img.save(image_path_)

def EXP_img2video(args,pipe, data_lists, input_dir, mask_dir, output_dir, start_index, end_index, diffusion_prompt):
    image_path = os.path.join(output_dir, os.path.basename(data_lists[start_index]['Image File']))
    image = load_image(image_path)
    image = image.resize((1024, 576))
    speed = data_lists[start_index]['Vehicle Speed']
    speed = speed * 2
    generator = torch.manual_seed(42)
    frames = pipe(image, decode_chunk_size=8,motion_bucket_id=speed+100, generator=generator).frames[0]
    for cont in range(0, args.pre_series):
        image_path_ = os.path.join(output_dir, os.path.basename(data_lists[start_index + cont]['Image File']))
        frames[cont].save(image_path_)


############################
def get_results_dir(args):
    save_dir = os.path.join(args.data_file, "results", "original")
    save_dir_2 = os.path.join(args.data_file, "results", "follow_up")
    save_dir_1 = os.path.join(args.data_file, "results", "mask")
    return save_dir,save_dir_1,save_dir_2


def get_MT_results_dir(args):
    save_dir = os.path.join(args.data_file, "MT_results", "original")
    save_dir_2 = os.path.join(args.data_file, "MT_results", "follow_up")
    save_dir_1 = os.path.join(args.data_file, "MT_results", "mask")
    save_dir_3 = os.path.join(args.data_file, "MT_results", "onroad_mask")
    save_dir_4 = os.path.join(args.data_file, "MT_results", "roadside_mask")
    return save_dir,save_dir_1,save_dir_2,save_dir_3,save_dir_4

def MT(args,MR):
    if args.MT_process==False:
        return 0
    if args.new_test_data==True:
        find_test_images(args)
    process_type = MR["type"]
    diffusion_prompt = MR["diffusion_prompt"]
    save_dir, save_dir_1, save_dir_2, save_dir_3, save_dir_4 = get_MT_results_dir(args)
    test_save_path = os.path.join(save_dir_2, str(MR["idx_new"]))
    if process_type=="Pix2Pix":
        MT_Pix2Pix(args, save_dir, test_save_path, diffusion_prompt, strength=1)
        return 0
    if process_type=="add_object":
        MT_add_object(args, save_dir, save_dir_1, save_dir_4, test_save_path, diffusion_prompt, strength=1)
    if process_type == "diffusion":
        diffusion_prompt=diffusion_prompt+"A wide road on the left side of the image, occupying the left third of the frame, first-person perspective as if from a car, clear visibility of the road's right edge, daytime, photorealistic, high detail,establishing shot,presented in full,untruncated"
        MT_diffusion(args, save_dir, save_dir_1, test_save_path, diffusion_prompt, strength=1)
    if args.pre_series>1:
        torch.cuda.empty_cache()
        vista(args,test_save_path)



def MT_ADS(args,MR):
    save_dir, save_dir_1, save_dir_2, save_dir_3, save_dir_4 = get_MT_results_dir(args)
    test_save_path = os.path.join(save_dir_2, str(MR["idx_new"]))
    Test_results = test_ADS.test_ads(args,save_dir,test_save_path)
    data_list = []
    for i in range(len(Test_results)):
        violation_tests,total_tests=test_ADS.get_violation(Test_results[i]["original"],Test_results[i]["follow_up"],MR["maneuver"])
        if i==0:
            data_str = f"""diffusion_prompt:{MR["diffusion_prompt"]},maneuver:{MR["maneuver"]},total:{total_tests},\n"""+f"""model:{Test_results[i]["model"]},violation:{violation_tests},"""
        else:
            data_str = data_str+f"""model:{Test_results[i]["model"]},violation:{violation_tests},"""
        """
        data={
            "model": Test_results[i]["model"],
            "maneuver": MR["maneuver"],
            "violation":violation_tests,
            "total":total_tests,
            "idx":MR["idx_new"],
            "diffusion_prompt":MR["diffusion_prompt"],
            "MR":MR["MR"]
        }
        data_list.append(data)
        """
    return data_str


def MT_diffusion(args,input_dir,mask,output_dir,diffusion_prompt,strength=1):
    shutil.rmtree(output_dir, ignore_errors=True)
    data_process.Check_file(output_dir)
    inpainter = FluxInpainting()
    with open(os.path.join(args.data_file, "MT_results","info.json"), 'r', encoding='utf-8') as file:
        data_lists = json.load(file)
    for idx in range(0, len(data_lists),args.pre_series):
        image_path = os.path.join(input_dir,os.path.basename(data_lists[idx]['Image File']))
        mask_path = os.path.join(mask, os.path.basename(data_lists[idx]['Image File']))
        
        image = inpainter(image_path,mask_path,prompt=diffusion_prompt)
        image.save(os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File'])))

def MT_add_object(args,input_dir,mask,road_side_mask,output_dir,diffusion_prompt,strength=0.3):
    shutil.rmtree(output_dir, ignore_errors=True)
    source_images = os.path.join(args.data_file,"MRs","image",f"{diffusion_prompt}.png")
    data_process.Check_file(output_dir)
    with open(os.path.join(args.data_file, "MT_results","info.json"), 'r', encoding='utf-8') as file:
        data_lists = json.load(file)
        for idx in range(0, len(data_lists),args.pre_series):
            image_path = os.path.join(input_dir,os.path.basename(data_lists[idx]['Image File']))
            mask_path = os.path.join(mask, os.path.basename(data_lists[idx]['Image File']))
            road_side_mask_path = os.path.join(road_side_mask, os.path.basename(data_lists[idx]['Image File']))
            image = Image.open(image_path)
            image = process_and_paste_image(image, mask_path, road_side_mask_path, source_images, scale_factor=strength)
            image.save(os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File'])))





def process_and_paste_image(original_image, mask_, mask_obj_, reference_image_,scale_factor=0.9):
    mask = cv2.imread(mask_)
    mask = mask[:,:, 0]
    mask_obj = cv2.imread(mask_obj_)
    mask_obj = mask_obj[:,:,0]
    reference_image = Image.open(reference_image_)
    point = Add_object.process_masks(mask, mask_obj)
    width = Add_object.integrated_road_width(mask, mask_obj, point)
    cropped_added_image, cropped_added_mask = Add_object.create_mask(reference_image)
    original_width, original_height = cropped_added_image.size
    scale_factor = 0.3
    new_width = int(width * scale_factor)
    new_height = int(original_height * (new_width / original_width))
    resized_image = cropped_added_image.resize((new_width, new_height), Image.LANCZOS)
    resized_mask = cropped_added_mask.resize((new_width, new_height), Image.LANCZOS)
    # 使用提供的粘贴点
    paste_y, paste_x = point
    paste_x = paste_x
    paste_y = paste_y - new_height
    image = original_image.copy()
    image_width, image_height = original_image.size

    ROAD_ID = 0
    road_mask = (mask == ROAD_ID).astype(np.uint8) * 255
    # 从当前行开始向上搜索
    for y in range(paste_y, -1, -1):
        road_pixels = np.where(road_mask[y] == 255)[0]
        if len(road_pixels) > 0:
            # 找到最右边的道路像素
            paste_x = road_pixels[-1]+new_width//8
            paste_y = y
            break

    if paste_x < 0:
        paste_x = 0
    elif paste_x + new_width > image_width:
        paste_x = image_width - new_width

    if paste_y < 0:
        paste_y = 0
    elif paste_y + new_height > image_height:
        paste_y = image_height - new_height
    image.paste(resized_image, (paste_x, paste_y), mask=resized_mask)
    return image

def MT_Pix2Pix(args,input_dir,output_dir,diffusion_prompt,strength=1):
    pix2pix = Pix2PixProcessor(strength)
    shutil.rmtree(output_dir, ignore_errors=True)
    data_process.Check_file(output_dir)
    with open(os.path.join(args.data_file, "MT_results","info.json"), 'r', encoding='utf-8') as file:
        data_lists = json.load(file)
    if args.RIEF == 1:
        for idx in range(0, len(data_lists)):
            if idx %2==0 or idx==len(data_lists)-1:
                image_path = os.path.join(input_dir,os.path.basename(data_lists[idx]['Image File']))
                image = Image.open(image_path)
                processed_image = pix2pix(diffusion_prompt, image)
                processed_image.save(os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File'])))
        for idx in range(0, len(data_lists)):
            if idx %2==1 and idx!=len(data_lists)-1:
                image_path = os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File']))
                image_path_0 = os.path.join(output_dir, os.path.basename(data_lists[idx - 1]['Image File']))
                image_path_1 = os.path.join(output_dir, os.path.basename(data_lists[idx + 1]['Image File']))
                fig_new = interpolate(image_path_0, image_path_1)
                cv2.imwrite(image_path, fig_new)
    else:
        for idx in range(0, len(data_lists)):
            image_path = os.path.join(input_dir,os.path.basename(data_lists[idx]['Image File']))
            image = Image.open(image_path)
            processed_image = pix2pix(diffusion_prompt, image)
            processed_image.save(os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File'])))

def vista(args,output_dir):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # 将当前目录（ECCV2022）的父目录（Diffusion_tools）添加到 Python 路径
    parent_dir = os.path.dirname(current_dir)
    sys.path.append(parent_dir)
    sys.path.append(os.path.join(parent_dir, "Other_tools", "Vista"))
    from my_test import VideoGenerator
    sys.stdout = open(os.devnull, 'w')

    import torchvision.transforms as transforms
    to_pil = transforms.ToPILImage()
    vista_gen = VideoGenerator()
    with open(os.path.join(args.data_file, "MT_results","info.json"), 'r', encoding='utf-8') as file:
        data_lists = json.load(file)
    for idx in range(0, len(data_lists), args.pre_series):
        image_path = os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File']))
        speed = data_lists[idx]['Vehicle Speed']
        generated_images = vista_gen(image_path,speed)
        if args.pre_series>25:
            print("should set the n_rounds and n_frames to generate more images")
            break
        for cont in range(0,args.pre_series):
            image_path_ = os.path.join(output_dir, os.path.basename(data_lists[idx+cont]['Image File']))
            img = to_pil(generated_images[cont])
            img.save(image_path_)


class Pix2PixProcessor:
    def __init__(self,strength):
        self.DeepTest_methods = ["DeepTest_" + effect for effect in get_corruption_names()]
        model_id = "timbrooks/instruct-pix2pix"
        self.pipe = StableDiffusionInstructPix2PixPipeline.from_pretrained(
            model_id,
            torch_dtype=torch.float16,
            safety_checker=None
        )
        self.strength = strength
        self.pipe.to("cuda")
        seed=42
        self.generator = torch.manual_seed(seed)
        self.pipe.scheduler = EulerAncestralDiscreteScheduler.from_config(self.pipe.scheduler.config)
    def __call__(self, diffusion_prompt, image):
        # Resize image
        width, height = image.size
        factor = 512 / max(width, height)
        factor = math.ceil(min(width, height) * factor / 64) * 64 / min(width, height)
        width = int((width * factor) // 64) * 64
        height = int((height * factor) // 64) * 64
        image = ImageOps.fit(image, (width, height), method=Image.Resampling.LANCZOS)

        if diffusion_prompt in self.DeepTest_methods:
            diffusion_prompt = diffusion_prompt[9:]
            np_image = np.array(image)
            temp_image = corrupt(np_image, corruption_name=diffusion_prompt, severity=self.strength)
            return Image.fromarray(temp_image)
        else:
            prompt = f"""what would it look like if it were {diffusion_prompt}?"""
            images = self.pipe(
                prompt,
                image=image,
                guidance_scale=self.strength+6.5,
                num_inference_steps=15,
                image_guidance_scale=1.5,
                generator = self.generator,
                disable_progress_bar = True  #
            ).images
            return images[0]


def RIEF(args,output_dir):
    from engine.RIFE import interpolate
    with open(os.path.join(args.data_file, "MT_results","info.json"), 'r', encoding='utf-8') as file:
        data_lists = json.load(file)
    for idx in range(0, len(data_lists)):
        if idx%2==1:
            image_path = os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File']))
            image_path_0 = os.path.join(output_dir, os.path.basename(data_lists[idx-1]['Image File']))
            image_path_1 = os.path.join(output_dir, os.path.basename(data_lists[idx + 1]['Image File']))
            fig_new = interpolate(image_path_0, image_path_1)
            cv2.imwrite(image_path, fig_new)

def resize_image_dimensions(original_resolution_wh, maximum_dimension=1024):
    width, height = original_resolution_wh
    if width > height:
        scaling_factor = maximum_dimension / width
    else:
        scaling_factor = maximum_dimension / height
    new_width = int(width * scaling_factor)
    new_height = int(height * scaling_factor)
    new_width = new_width - (new_width % 32)
    new_height = new_height - (new_height % 32)
    return new_width, new_height


class FluxInpainting:
    def __init__(self, model_path="black-forest-labs/FLUX.1-schnell", image_size=1024):
        os.environ['PYTORCH_CUDA_ALLOC_CONF'] = 'expandable_segments:True'
        self.DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
        self.IMAGE_SIZE = image_size

        self.pipe = FluxInpaintPipeline.from_pretrained(model_path,torch_dtype=torch.bfloat16)
        self.pipe.enable_model_cpu_offload()

        gc.collect()
        torch.cuda.empty_cache()
        torch.cuda.memory_summary(device=None, abbreviated=False)

    def create_mask(self, image_path,mask_path,type = "less", bottom_white_percentage=0.6):
        image = Image.open(image_path)
        width, height = image.size
        mask = cv2.imread(mask_path)
        height, width = mask.shape[:2]
        mask_1 = np.ones((height, width), dtype=np.uint8) * 255
        if type=="less":
            elements_to_keep = [2, 10]  # 2: building, 8: vegetation, 10: sky [2,8,10]8,
            kernel_size = 40
        if type=="more":
            elements_to_keep = [1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14, 15, 16, 17, 18]
            kernel_size = 20
        if type=="little":
            elements_to_keep = [14,15]
            kernel_size = 10
        #elements_to_keep = [1, 2, 3, 4, 5, 6, 7,  9, 11, 12, 13, 14, 15, 16, 17, 18]
        if len(mask.shape) == 3:
            for element in elements_to_keep:
                mask_1[np.all(mask == element, axis=-1)] = 0
        else:
            for element in elements_to_keep:
                mask_1[mask == element] = 0
        inverted_mask = cv2.bitwise_not(mask_1)


        kernel = np.ones((kernel_size, kernel_size), np.uint8)
        # 执行腐蚀操作
        iterations = 1
        eroded_mask_1 = cv2.erode(inverted_mask, kernel, iterations=iterations)
        dilated_mask = cv2.bitwise_not(eroded_mask_1)

        mask = Image.fromarray(dilated_mask)
        return image, mask

    def __call__(self, image_path, mask_path, prompt,type="less", seed=42, strength=0.85, num_inference_steps=40):
        seed = random.randint(0, 1000000)
        image,mask = self.create_mask(image_path,mask_path,type)
        width, height = resize_image_dimensions(image.size, self.IMAGE_SIZE)
       # width, height = 1024,1024
        if type =="little":
            width, height = 256, 256
        resized_image = image.resize((width, height), Image.LANCZOS)
        resized_mask = mask.resize((width, height), Image.LANCZOS)
        generator = torch.Generator("cpu").manual_seed(seed)
        prompt ="(the road facing directly towards the viewer),photorealistic,high detail,"+prompt#prompt+",photorealistic, (the road facing directly towards the viewer),high detail"#,(The viewpoint must drive on the road, and in the center lane of the road),
        with torch.no_grad():
            result = self.pipe(
                prompt=prompt,
                image=resized_image,
                mask_image=resized_mask,
                width=width,
                height=height,
                strength=strength,
                max_sequence_length=256,
                generator=generator,
                num_inference_steps=num_inference_steps
            ).images[0]

        return result


def MT_1(args,type,MRs):
    type_specific_MRs = [mr for mr in MRs if mr["type"] == type]

    tem=1
    if tem ==0:
        inpainter = FluxInpainting()
        for i in range(len(type_specific_MRs)):
            MR = type_specific_MRs[i]
            process_type = MR["type"]
            diffusion_prompt = MR["diffusion_prompt"]
            save_dir, save_dir_1, save_dir_2, save_dir_3, save_dir_4 = get_MT_results_dir(args)
            test_save_path = os.path.join(save_dir_2, str(MR["idx_new"]))
            if process_type=="Pix2Pix":
                MT_Pix2Pix(args, save_dir, test_save_path, diffusion_prompt, strength=1)
            if process_type == "diffusion":
                diffusion_prompt=diffusion_prompt+", no raised median, no construction debris,A wide road on the left side of the image, POV: Looking at the right side of the road, first-person perspective as if from a car, presented in full, untruncated"
                MT_diffusion_1(args,inpainter, save_dir, save_dir_1, test_save_path, diffusion_prompt, strength=1)
    if args.pre_series>1:
        for i in range(len(type_specific_MRs)):
            MR = type_specific_MRs[i]
            save_dir, save_dir_1, save_dir_2, save_dir_3, save_dir_4 = get_MT_results_dir(args)
            test_save_path = os.path.join(save_dir_2, str(MR["idx_new"]))
            torch.cuda.empty_cache()
            current_dir = os.path.dirname(os.path.abspath(__file__))
            # 将当前目录（ECCV2022）的父目录（Diffusion_tools）添加到 Python 路径
            parent_dir = os.path.dirname(current_dir)
            sys.path.append(parent_dir)
            sys.path.append(os.path.join(parent_dir, "Other_tools", "Vista"))
            from my_test import VideoGenerator
            sys.stdout = open(os.devnull, 'w')

            import torchvision.transforms as transforms

            vista_gen = VideoGenerator()

            vista_1(args,vista_gen,test_save_path)

def MT_2(args,type,MRs):
    inpainter = FluxInpainting()
    pro_lists = [
        "Vehicle approaching from right and pedestrian crossing, the position of these objects in unpaved road intersecting paved road",
        "Truck making a right turn while bicyclist approaches from left, the position of these objects in busy urban intersection",
        "School bus stopping and children crossing street, the position of these objects in suburban road near school zone",
        "Motorcycle lane splitting and car changing lanes, the position of these objects in congested multi-lane highway",
        "Bus at designated stop and passengers rushing to board, the position of these objects in dedicated bus lane on city street",
        "Bicycle making a left turn and pedestrian using crosswalk, the position of these objects in roundabout with multiple exits",
        "Emergency vehicle with sirens and cars pulling over, the position of these objects in middle lane of three-lane road",
        "Tractor moving slowly and sports car overtaking, the position of these objects on rural single-lane road with no shoulder",
        "Car parallel parking and delivery truck double parked, the position of these objects on busy commercial street",
        "Train crossing and cars waiting at barriers, the position of these objects on tracks intersecting major highway",
        "Taxi picking up passenger and rideshare vehicle dropping off, the position of these objects at airport terminal zone",
        "Delivery van unloading and construction worker crossing, the position of these objects on narrow one-way street",
        "Scooter weaving through traffic and pedestrian jaywalking, the position of these objects at busy intersection",
        "Horse-drawn carriage touring and tourists taking photos, the position of these objects on cobblestone street in historic district",
        "Garbage truck stopping and children waiting for school bus, the position of these objects on residential street",
        "Police car pursuing speeding vehicle, the position of these objects on divided highway with median",
        "School bus extending stop sign and cyclist approaching from behind, the position of these objects on suburban road",
        "Ambulance navigating through intersection and pedestrians scrambling to clear path, the position of these objects in city center",
        "Snowplow clearing road and stranded vehicle on shoulder, the position of these objects on winding mountain road during blizzard",
        "Motorcycle cop monitoring speed and truck weighing station ahead, the position of these objects on highway with speed trap",
        "Trolley on fixed track and tourists crossing carelessly, the position of these objects on steep hill in coastal city",
        "Food truck at event and long queue of customers, the position of these objects in public park area with pedestrian paths",
        "Cement mixer turning into construction site and flagman directing traffic, the position of these objects at entrance of project",
        "Limousine picking up VIP and paparazzi on sidewalk, the position of these objects in front of luxury hotel",
        "Fire engine responding to call and cars struggling to make way, the position of these objects on narrow bridge",
        "Electric car charging and gasoline car waiting for spot, the position of these objects in parking lot of shopping center",
        "Tanker truck negotiating sharp turn and wildlife crossing sign ahead, the position of these objects on cliffside road",
        "Street sweeper cleaning and bicycle navigating around, the position of these objects on wide boulevard with bike lanes",
        "Tow truck assisting broken-down vehicle and rubbernecking causing slowdown, the position of these objects on busy freeway",
        "Self-driving car and skateboarder at crosswalk, the position of these objects in mixed residential and commercial area"
    ]

    temp = 30
    for i in range(temp):
        process_type = "diffusion"
        diffusion_prompt = pro_lists[i]+", no raised median, no construction debris,A wide road on the left side of the image, POV: Looking at the right side of the road, first-person perspective as if from a car, presented in full, untruncated"
        save_dir, save_dir_1, save_dir_2, save_dir_3, save_dir_4 = get_MT_results_dir(args)
        test_save_path = os.path.join(save_dir_2, str(i+100))
        if process_type=="Pix2Pix":
            MT_Pix2Pix(args, save_dir, test_save_path, diffusion_prompt, strength=1)
        if process_type == "diffusion":
            diffusion_prompt=diffusion_prompt
            MT_diffusion_1(args,inpainter, save_dir, save_dir_1, test_save_path, diffusion_prompt, strength=1)
    torch.cuda.empty_cache()
    pipe = StableDiffusionXLImg2ImgPipeline.from_pretrained(
        "stabilityai/stable-diffusion-xl-refiner-1.0", torch_dtype=torch.float16, variant="fp16", use_safetensors=True
    )
    pipe = pipe.to("cuda")
    for i in range(temp):
        save_dir, save_dir_1, save_dir_2, save_dir_3, save_dir_4 = get_MT_results_dir(args)
        test_save_path = os.path.join(save_dir_2, str(i+100))

        with open(os.path.join(args.data_file, "MT_results", "info.json"), 'r', encoding='utf-8') as file:
            data_lists = json.load(file)
        for idx in range(0, len(data_lists), args.pre_series):
            image_path = os.path.join(test_save_path, os.path.basename(data_lists[idx]['Image File']))
            image = Image.open(image_path)
            image.save(os.path.join(test_save_path, "_"+os.path.basename(data_lists[idx]['Image File'])))
            init_image = load_image(image_path).convert("RGB")
            prompt = "Busy traffic scene, muted colors, detailed, 8k"
            image = pipe(prompt, image=init_image).images
            image_path = os.path.join(test_save_path, os.path.basename(data_lists[idx]['Image File']))
            image[0].save(image_path)

    if args.pre_series>1:
        for i in range(temp):
            save_dir, save_dir_1, save_dir_2, save_dir_3, save_dir_4 = get_MT_results_dir(args)
            test_save_path = os.path.join(save_dir_2, str(i+100))
            torch.cuda.empty_cache()
            current_dir = os.path.dirname(os.path.abspath(__file__))
            # 将当前目录（ECCV2022）的父目录（Diffusion_tools）添加到 Python 路径
            parent_dir = os.path.dirname(current_dir)
            sys.path.append(parent_dir)
            sys.path.append(os.path.join(parent_dir, "Other_tools", "Vista"))
            from my_test import VideoGenerator
            sys.stdout = open(os.devnull, 'w')


            to_pil = transforms.ToPILImage()
            vista_gen = VideoGenerator()

            vista_1(args,vista_gen,test_save_path)

def MT_diffusion_1(args,inpainter, input_dir, mask, output_dir, diffusion_prompt, strength=1):
    shutil.rmtree(output_dir, ignore_errors=True)
    data_process.Check_file(output_dir)

    with open(os.path.join(args.data_file, "MT_results", "info.json"), 'r', encoding='utf-8') as file:
        data_lists = json.load(file)
    for idx in range(0, len(data_lists), args.pre_series):
        image_path = os.path.join(input_dir, os.path.basename(data_lists[idx]['Image File']))
        mask_path = os.path.join(mask, os.path.basename(data_lists[idx]['Image File']))

        image = inpainter(image_path, mask_path, prompt=diffusion_prompt)
        image.save(os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File'])))

def MT_refiner_1(args,inpainter, input_dir, mask, output_dir, diffusion_prompt, strength=1):



        image = inpainter(image_path, mask_path, prompt=diffusion_prompt)
        image.save(os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File'])))

def vista_1(args,vista_gen,output_dir):

    with open(os.path.join(args.data_file, "MT_results","info.json"), 'r', encoding='utf-8') as file:
        data_lists = json.load(file)
    for idx in range(0, len(data_lists), args.pre_series):
        image_path = os.path.join(output_dir, os.path.basename(data_lists[idx]['Image File']))
        speed = data_lists[idx]['Vehicle Speed']
        speed = speed*2
        generated_images = vista_gen(image_path,speed)
        if args.pre_series>25:
            print("should set the n_rounds and n_frames to generate more images")
            break
        for cont in range(0,args.pre_series):
            image_path_ = os.path.join(output_dir, os.path.basename(data_lists[idx+cont]['Image File']))
            to_pil = transforms.ToPILImage()
            img = to_pil(generated_images[cont])
            img.save(image_path_)




